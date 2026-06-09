#!/usr/bin/env python3
from __future__ import annotations

import re
import uuid
import json
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
MAPPING_XLSX = ROOT / "database" / "source" / "CookBuddy_Mapping_Review.xlsx"
OUTPUT_SQL = ROOT / "database" / "seed.sql"
OUTPUT_JSON = ROOT / "database" / "generated" / "recipes.json"

NAMESPACE = uuid.UUID("e5a62b31-2844-49ba-b32f-d5f06c1e41f4")

IMAGE_BY_MEAL = {
    "breakfast": "/assets/images/dishes/breakfast-default.png",
    "lunch": "/assets/images/dishes/lunch-default.png",
    "dinner": "/assets/images/dishes/dinner-default.png",
    "snack": "/assets/images/snacks/snacks-default.png",
}

IMAGE_MATCHERS = [
    ("chicken 65", "/assets/images/snacks/chicken-65.png"),
    ("mirchi bajji", "/assets/images/snacks/mirchi-bajji.png"),
    ("pakora", "/assets/images/snacks/pakora.png"),
    ("samosa", "/assets/images/snacks/samosa.png"),
    ("dosa", "/assets/images/dishes/dosa.png"),
    ("idli", "/assets/images/dishes/idli.png"),
    ("pongal", "/assets/images/dishes/pongal.png"),
    ("poha", "/assets/images/dishes/poha.png"),
    ("upma", "/assets/images/dishes/upma.png"),
    ("aloo paratha", "/assets/images/dishes/aloo-paratha.png"),
    ("paratha", "/assets/images/dishes/paratha.png"),
    ("bread omelette", "/assets/images/dishes/bread-omelette.png"),
    ("rajma", "/assets/images/dishes/rajma-chawal.png"),
    ("sambar rice", "/assets/images/dishes/sambar-rice.png"),
    ("lemon rice", "/assets/images/dishes/lemon-rice.png"),
    ("khichdi", "/assets/images/dishes/khichdi.png"),
    ("fish curry", "/assets/images/dishes/fish-curry-rice.png"),
    ("chicken curry", "/assets/images/dishes/chicken-curry-rice.png"),
    ("curd rice", "/assets/images/dishes/curd-rice.png"),
    ("dal rice", "/assets/images/dishes/dal-rice.png"),
    ("pulao", "/assets/images/dishes/pulao.png"),
    ("egg curry", "/assets/images/dishes/egg-curry.png"),
    ("paneer", "/assets/images/dishes/paneer-curry.png"),
    ("rasam rice", "/assets/images/dishes/rasam-rice.png"),
    ("soup", "/assets/images/dishes/soup-bowls.png"),
    ("biryani", "/assets/images/dishes/biryani.png"),
    ("ladoo", "/assets/images/desserts/laddoo.png"),
    ("sweet", "/assets/images/desserts/dessert-default.png"),
]

PROTEINS = {
    "egg",
    "chicken",
    "fish",
    "pork",
    "paneer",
    "chana",
    "rajma",
    "dal",
    "toor-dal",
    "moong-dal",
    "urad-dal",
    "besan",
}

GRAINS = {
    "rice",
    "idli-rice",
    "wheat",
    "rava",
    "poha",
    "oats",
    "ragi",
    "sabudana",
    "noodles",
    "bread",
}

DAIRY = {"curd", "milk", "ghee", "khoya"}


def stable_uuid(kind: str, value: str) -> str:
    return str(uuid.uuid5(NAMESPACE, f"{kind}:{value}"))


def sql(value) -> str:
    if value is None:
        return "NULL"
    if isinstance(value, bool):
        return "TRUE" if value else "FALSE"
    if isinstance(value, (int, float)):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def pg_array(values: list[str]) -> str:
    return "ARRAY[" + ", ".join(sql(value) for value in values) + "]"


def rows_from_sheet(workbook, sheet_name: str) -> list[dict[str, object]]:
    sheet = workbook[sheet_name]
    headers = [cell.value for cell in sheet[2]]
    rows = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        if not any(value is not None for value in row):
            continue
        item = {headers[index]: value for index, value in enumerate(row) if index < len(headers) and headers[index]}
        rows.append(item)
    return rows


def category_for(ingredient_id: str) -> str:
    if ingredient_id in PROTEINS:
        return "protein"
    if ingredient_id in GRAINS:
        return "grain"
    if ingredient_id in DAIRY:
        return "dairy"
    if ingredient_id in {"apple", "banana", "lemon", "coconut", "peanut", "dry-fruits"}:
        return "fruit"
    if ingredient_id in {"pepper", "tamarind"}:
        return "spice"
    return "vegetable"


def default_unit_for(ingredient_id: str) -> str:
    if ingredient_id in {"egg", "banana", "apple", "lemon", "onion", "tomato", "potato"}:
        return "pcs"
    if ingredient_id in {"curd", "milk"}:
        return "cup"
    if ingredient_id in {"ghee", "pepper"}:
        return "tsp"
    return "g"


def meal_from_tags(tags: list[str]) -> str:
    for meal in ("breakfast", "lunch", "dinner", "snack"):
        if meal in tags:
            return meal
    return "dinner"


def image_for_recipe(dish_name: str, tags: list[str], meal: str) -> str:
    text = f"{dish_name} {' '.join(tags)}".lower()
    for term, image_path in IMAGE_MATCHERS:
        if term in text:
            return image_path
    return IMAGE_BY_MEAL[meal]


def title_case(text: str) -> str:
    return re.sub(r"\s+", " ", text).strip().title()


def main_ingredient_names(ingredients: list[dict[str, object]]) -> list[str]:
    mains = [str(row["ingredient_name"]) for row in ingredients if row.get("is_main")]
    return mains or [str(row["ingredient_name"]) for row in ingredients[:2]]


def load_mapping():
    workbook = load_workbook(MAPPING_XLSX, data_only=True)
    recipe_meta = {
        str(row["recipe_id"]): row
        for row in rows_from_sheet(workbook, "recipes")
    } if "recipes" in workbook.sheetnames else {}
    recipe_ingredients = rows_from_sheet(workbook, "recipe_ingredients")
    recipe_tags = rows_from_sheet(workbook, "recipe_tags")
    recipe_pairings = rows_from_sheet(workbook, "recipe_pairings")
    ingredient_aliases = rows_from_sheet(workbook, "ingredient_aliases")

    ingredients: dict[str, str] = {}
    recipe_names: dict[str, str] = {}
    ingredients_by_recipe: dict[str, list[dict[str, object]]] = defaultdict(list)
    tags_by_recipe: dict[str, list[str]] = defaultdict(list)

    for row in recipe_ingredients:
        recipe_id = str(row["recipe_id"])
        ingredient_id = str(row["ingredient_id"])
        recipe_names[recipe_id] = str(row["dish_name"])
        ingredients[ingredient_id] = str(row["ingredient_name"])
        ingredients_by_recipe[recipe_id].append(row)

    for row in recipe_tags:
        tags_by_recipe[str(row["recipe_id"])].append(str(row["tag"]).lower())
        recipe_names.setdefault(str(row["recipe_id"]), str(row["dish_name"]))

    return {
        "recipe_ingredients": recipe_ingredients,
        "recipe_tags": recipe_tags,
        "recipe_pairings": recipe_pairings,
        "ingredient_aliases": ingredient_aliases,
        "ingredients": ingredients,
        "recipe_names": recipe_names,
        "ingredients_by_recipe": ingredients_by_recipe,
        "tags_by_recipe": tags_by_recipe,
        "recipe_meta": recipe_meta,
    }


def diet_from_meta(meta: dict[str, object], tags: list[str]) -> str:
    value = str(meta.get("veg_type") or "").strip().lower()
    if value in {"non-vegetarian", "vegetarian", "eggetarian"}:
        return value
    return "non-vegetarian" if "non-vegetarian" in tags else "vegetarian"


def difficulty_from_tags(tags: list[str]) -> str:
    return "easy" if "quick" in tags or "quick-meal" in tags or "comfort" in tags or "comfort-food" in tags else "medium"


def time_from_meta(meta: dict[str, object]) -> tuple[int, int]:
    total = int(meta.get("prep_time_mins") or 30)
    prep = min(10, max(5, total // 3))
    return prep, max(10, total - prep)


def number_or_none(value) -> float | int | None:
    if value in (None, ""):
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    return int(numeric) if numeric.is_integer() else numeric


def int_score(value, default: int | None = None) -> int | None:
    numeric = number_or_none(value)
    if numeric is None:
        return default
    return max(1, min(10, int(round(float(numeric)))))


def secondary_profile(dish_name: str, tags: list[str], ingredients: list[dict[str, object]], meta: dict[str, object]) -> dict[str, object]:
    text = f"{dish_name} {' '.join(tags)} {' '.join(str(row.get('ingredient_name') or '') for row in ingredients)}".lower()
    tag_set = set(tags)
    total_time = int(meta.get("prep_time_mins") or 30)
    meal = meal_from_tags(tags)
    is_quick = total_time <= 25 or "quick-meal" in tag_set or "quick" in tag_set
    is_comfort = "comfort-food" in tag_set or "comfort" in tag_set
    is_rainy = "rainy-day" in tag_set
    is_protein = "high-protein" in tag_set or any(word in text for word in ["egg", "chicken", "fish", "mutton", "pork", "paneer", "dal", "chana", "rajma"])
    is_cooling = any(word in text for word in ["curd", "buttermilk", "lassi", "cucumber", "salad", "fruit", "rose milk"])
    is_light = any(word in text for word in ["rasam", "curd rice", "soup", "salad", "porridge", "puree", "idli", "poha"]) or total_time <= 20
    is_one_pot = any(word in text for word in ["khichdi", "rice", "pulao", "biryani", "bisibelebath", "sambar rice", "curd rice", "rasam rice", "soup", "stew"])
    is_snack = meal == "snack" or "snack" in tag_set or "snacks" in tag_set
    is_late_night = is_snack or is_quick or any(word in text for word in ["maggi", "momos", "bread omelette", "chai", "poha", "upma", "khichdi"])
    effort_score = int_score(meta.get("effort_score"))
    if effort_score is None:
        effort_score = 2 if total_time <= 15 else 3 if total_time <= 25 else 5 if total_time <= 40 else 8
    comfort_score = int_score(meta.get("comfort_score"), 5)
    if is_comfort:
        comfort_score = max(comfort_score, 8)
    if is_rainy:
        comfort_score = max(comfort_score, 7)
    energy_score = int_score(meta.get("energy_score"))
    if energy_score is None:
        energy_score = 3 if is_light or is_quick else 6 if is_protein else 5
    return {
        "lateNight": bool(meta.get("late_night")) or is_late_night,
        "lowEffort": bool(meta.get("low_effort")) or effort_score <= 3 or is_quick,
        "sickDay": bool(meta.get("sick_day")) or any(word in text for word in ["rasam", "khichdi", "soup", "porridge", "curd rice"]),
        "budgetFriendly": bool(meta.get("budget_friendly")) or any(word in text for word in ["rice", "dal", "poha", "upma", "idli", "chilla", "corn", "banana", "potato", "egg"]),
        "summerCooling": bool(meta.get("summer_cooling")) or is_cooling,
        "lightMeal": bool(meta.get("light_meal")) or is_light,
        "onePot": bool(meta.get("one_pot")) or is_one_pot,
        "minimalCleanup": bool(meta.get("minimal_cleanup")) or is_quick or is_one_pot,
        "studySnack": bool(meta.get("study_snack")) or (is_snack and is_quick),
        "weekendSpecial": bool(meta.get("weekend_special")) or total_time >= 45 or "festival" in tag_set or any(word in text for word in ["biryani", "butter chicken", "chettinad", "laal maas", "gongura mutton"]),
        "effortScore": effort_score,
        "comfortScore": comfort_score,
        "energyScore": energy_score,
    }


def emotional_profile(meta: dict[str, object]) -> dict[str, object]:
    fields = {
        "emotionalState": "emotional_state",
        "soulFoodScore": "soul_food_score",
        "rainyDayScore": "rainy_day_score",
        "proteinScore": "protein_score",
        "nostalgiaScore": "nostalgia_score",
        "stomachFeel": "stomach_feel",
        "bestTime": "best_time",
        "bestDay": "best_day",
        "tomoLine": "tomo_line",
        "homeStyleScore": "home_style_score",
    }
    profile: dict[str, object] = {}
    for output_key, source_key in fields.items():
        value = meta.get(source_key)
        if source_key.endswith("_score"):
            value = number_or_none(value)
        if value not in (None, ""):
            profile[output_key] = value
    return profile


def emotional_sql_values(meta: dict[str, object]) -> list[str]:
    return [
        sql(meta.get("emotional_state")),
        sql(int_score(meta.get("soul_food_score"))),
        sql(int_score(meta.get("rainy_day_score"))),
        sql(int_score(meta.get("protein_score"))),
        sql(int_score(meta.get("nostalgia_score"))),
        sql(meta.get("stomach_feel")),
        sql(meta.get("best_time")),
        sql(meta.get("best_day")),
        sql(meta.get("tomo_line")),
        sql(int_score(meta.get("home_style_score"))),
    ]


def sentence_list(values: list[str]) -> str:
    cleaned = [str(value).strip().lower() for value in values if str(value).strip()]
    if not cleaned:
        return "simple pantry staples"
    if len(cleaned) == 1:
        return cleaned[0]
    if len(cleaned) == 2:
        return f"{cleaned[0]} and {cleaned[1]}"
    return ", ".join(cleaned[:-1]) + f", and {cleaned[-1]}"


def describe_recipe(dish_name: str, mains: list[str], tags: list[str], meal: str, diet_type: str) -> str:
    tag_set = set(tags)
    ingredients = sentence_list(mains[:4])
    meal_label = {
        "breakfast": "morning",
        "lunch": "midday",
        "dinner": "dinner",
        "snack": "snack-time",
    }.get(meal, "home")

    if "spicy-food" in tag_set or "spicy" in tag_set:
        mood_line = "It brings a lively chilli warmth without losing that familiar home-cooked comfort."
    elif "rainy-day" in tag_set:
        mood_line = "It feels especially right on a rainy day, when you want something warm, crisp, or soothing."
    elif "high-protein" in tag_set or diet_type in {"non-vegetarian", "eggetarian"}:
        mood_line = "It is filling enough to anchor the meal while still feeling simple and practical."
    elif "quick-meal" in tag_set or "quick" in tag_set:
        mood_line = "It comes together quickly, so it works well when you want real food without a long cooking session."
    elif "comfort-food" in tag_set or "comfort" in tag_set:
        mood_line = "It has the soft, familiar comfort of food you would happily eat at home on a slower day."
    else:
        mood_line = "It keeps the flavours familiar, balanced, and easy to enjoy."

    return (
        f"{dish_name} is a home-style {meal_label} idea made with {ingredients}. "
        f"{mood_line} "
        "Serve it warm, adjust the spice to your mood, and pair it with a simple side if you want the plate to feel complete."
    )


def cooking_instructions(dish_name: str, mains: list[str]) -> list[str]:
    ingredients = sentence_list(mains[:4])
    return [
        f"Prep {ingredients} so everything is ready before the pan gets hot.",
        "Start with the usual home base of oil or ghee, onion, tomato, herbs, or spices as the dish needs.",
        f"Cook gently until the flavours come together, then finish the {dish_name} with salt, heat, and texture adjusted to taste.",
        "Serve warm while it still feels fresh and comforting.",
    ]


def recipe_payload(recipe_id: str, dish_name: str, ingredients_by_recipe, tags_by_recipe, recipe_meta) -> dict[str, object]:
    recipe_tags_for_item = tags_by_recipe[recipe_id]
    meta = recipe_meta.get(recipe_id, {})
    meal = meal_from_tags(recipe_tags_for_item)
    ingredients_for_item = ingredients_by_recipe[recipe_id]
    mains = main_ingredient_names(ingredients_for_item)
    diet_type = diet_from_meta(meta, recipe_tags_for_item)
    difficulty = difficulty_from_tags(recipe_tags_for_item)
    prep_time, cook_time = time_from_meta(meta)
    secondary = secondary_profile(dish_name, recipe_tags_for_item, ingredients_for_item, meta)
    return {
        "id": stable_uuid("recipe", recipe_id),
        "sourceId": recipe_id,
        "title": dish_name,
        "description": describe_recipe(dish_name, mains, recipe_tags_for_item, meal, diet_type),
        "prepTimeMinutes": prep_time,
        "cookTimeMinutes": cook_time,
        "servings": 2,
        "cuisine": "Indian",
        "dietType": diet_type,
        "difficulty": difficulty,
        "imageUrl": image_for_recipe(dish_name, recipe_tags_for_item, meal),
        "tags": recipe_tags_for_item,
        **secondary,
        **emotional_profile(meta),
        "instructions": cooking_instructions(dish_name, mains),
        "ingredients": [
            {
                "name": str(row["ingredient_name"]),
                "quantity": 1 if bool(row.get("is_main")) else 0.5,
                "unit": default_unit_for(str(row["ingredient_id"])),
                "role": str(row.get("role") or "required"),
                "isMain": bool(row.get("is_main")),
            }
            for row in ingredients_for_item
        ],
    }


def build_fallback_json(mapping) -> list[dict[str, object]]:
    return [
        recipe_payload(recipe_id, dish_name, mapping["ingredients_by_recipe"], mapping["tags_by_recipe"], mapping["recipe_meta"])
        for recipe_id, dish_name in sorted(mapping["recipe_names"].items())
    ]


def build_seed(mapping) -> str:
    recipe_ingredients = mapping["recipe_ingredients"]
    recipe_tags = mapping["recipe_tags"]
    recipe_pairings = mapping["recipe_pairings"]
    ingredient_aliases = mapping["ingredient_aliases"]
    ingredients = mapping["ingredients"]
    recipe_names = mapping["recipe_names"]
    ingredients_by_recipe = mapping["ingredients_by_recipe"]
    tags_by_recipe = mapping["tags_by_recipe"]
    recipe_meta = mapping["recipe_meta"]

    lines = [
        "-- Generated from database/source/CookBuddy_Mapping_Review.xlsx.",
        "-- Rebuild with: python3 scripts/generate_seed_from_mapping.py",
        "BEGIN;",
        "",
        "TRUNCATE recipe_ingredients, recipe_pairings, recipe_tags, ingredient_aliases, recipes, ingredients RESTART IDENTITY CASCADE;",
        "",
    ]

    lines.append("INSERT INTO ingredients (id, name, category, default_unit) VALUES")
    ingredient_values = []
    for ingredient_id, name in sorted(ingredients.items()):
        ingredient_values.append(
            f"  ({sql(stable_uuid('ingredient', ingredient_id))}, {sql(name)}, {sql(category_for(ingredient_id))}, {sql(default_unit_for(ingredient_id))})"
        )
    lines.append(",\n".join(ingredient_values) + "\nON CONFLICT (name) DO UPDATE SET category = EXCLUDED.category, default_unit = EXCLUDED.default_unit;")
    lines.append("")

    lines.append("INSERT INTO recipes (id, title, description, instructions, prep_time_minutes, cook_time_minutes, servings, cuisine, diet_type, difficulty, image_url, source_id, emotional_state, late_night, low_effort, sick_day, budget_friendly, summer_cooling, light_meal, one_pot, minimal_cleanup, study_snack, weekend_special, effort_score, comfort_score, energy_score, soul_food_score, rainy_day_score, protein_score, nostalgia_score, stomach_feel, best_time, best_day, tomo_line, home_style_score) VALUES")
    recipe_values = []
    for recipe_id, dish_name in sorted(recipe_names.items()):
        recipe_tags_for_item = tags_by_recipe[recipe_id]
        meta = recipe_meta.get(recipe_id, {})
        meal = meal_from_tags(recipe_tags_for_item)
        mains = main_ingredient_names(ingredients_by_recipe[recipe_id])
        diet_type = diet_from_meta(meta, recipe_tags_for_item)
        difficulty = difficulty_from_tags(recipe_tags_for_item)
        prep_time, cook_time = time_from_meta(meta)
        description = describe_recipe(dish_name, mains, recipe_tags_for_item, meal, diet_type)
        instructions = cooking_instructions(dish_name, mains)
        secondary = secondary_profile(dish_name, recipe_tags_for_item, ingredients_by_recipe[recipe_id], meta)
        recipe_values.append(
            "  ("
            + ", ".join(
                [
                    sql(stable_uuid("recipe", recipe_id)),
                    sql(dish_name),
                    sql(description),
                    pg_array(instructions),
                    str(prep_time),
                    str(cook_time),
                    "2",
                    sql("Indian"),
                    sql(diet_type),
                    sql(difficulty),
                    sql(image_for_recipe(dish_name, recipe_tags_for_item, meal)),
                    sql(recipe_id),
                    sql(meta.get("emotional_state")),
                    sql(secondary["lateNight"]),
                    sql(secondary["lowEffort"]),
                    sql(secondary["sickDay"]),
                    sql(secondary["budgetFriendly"]),
                    sql(secondary["summerCooling"]),
                    sql(secondary["lightMeal"]),
                    sql(secondary["onePot"]),
                    sql(secondary["minimalCleanup"]),
                    sql(secondary["studySnack"]),
                    sql(secondary["weekendSpecial"]),
                    sql(secondary["effortScore"]),
                    sql(secondary["comfortScore"]),
                    sql(secondary["energyScore"]),
                    *emotional_sql_values(meta)[1:],
                ]
            )
            + ")"
        )
    lines.append(",\n".join(recipe_values) + "\nON CONFLICT (source_id) DO UPDATE SET title = EXCLUDED.title, description = EXCLUDED.description, instructions = EXCLUDED.instructions, prep_time_minutes = EXCLUDED.prep_time_minutes, cook_time_minutes = EXCLUDED.cook_time_minutes, servings = EXCLUDED.servings, cuisine = EXCLUDED.cuisine, diet_type = EXCLUDED.diet_type, difficulty = EXCLUDED.difficulty, image_url = EXCLUDED.image_url, emotional_state = EXCLUDED.emotional_state, late_night = EXCLUDED.late_night, low_effort = EXCLUDED.low_effort, sick_day = EXCLUDED.sick_day, budget_friendly = EXCLUDED.budget_friendly, summer_cooling = EXCLUDED.summer_cooling, light_meal = EXCLUDED.light_meal, one_pot = EXCLUDED.one_pot, minimal_cleanup = EXCLUDED.minimal_cleanup, study_snack = EXCLUDED.study_snack, weekend_special = EXCLUDED.weekend_special, effort_score = EXCLUDED.effort_score, comfort_score = EXCLUDED.comfort_score, energy_score = EXCLUDED.energy_score, soul_food_score = EXCLUDED.soul_food_score, rainy_day_score = EXCLUDED.rainy_day_score, protein_score = EXCLUDED.protein_score, nostalgia_score = EXCLUDED.nostalgia_score, stomach_feel = EXCLUDED.stomach_feel, best_time = EXCLUDED.best_time, best_day = EXCLUDED.best_day, tomo_line = EXCLUDED.tomo_line, home_style_score = EXCLUDED.home_style_score;")
    lines.append("")

    lines.append("INSERT INTO recipe_ingredients (id, recipe_id, ingredient_id, quantity, unit, role, is_main, match_weight) VALUES")
    mapping_values = []
    for row in recipe_ingredients:
        ingredient_id = str(row["ingredient_id"])
        is_main = bool(row.get("is_main"))
        quantity = 1 if is_main else 0.5
        mapping_values.append(
            "  ("
            + ", ".join(
                [
                    sql(stable_uuid("recipe_ingredient", str(row["id"]))),
                    sql(stable_uuid("recipe", str(row["recipe_id"]))),
                    sql(stable_uuid("ingredient", ingredient_id)),
                    sql(quantity),
                    sql(default_unit_for(ingredient_id)),
                    sql(row.get("role") or "required"),
                    sql(is_main),
                    sql(row.get("match_weight") or (1 if is_main else 0.35)),
                ]
            )
            + ")"
        )
    lines.append(",\n".join(mapping_values) + "\nON CONFLICT (recipe_id, ingredient_id) DO UPDATE SET quantity = EXCLUDED.quantity, unit = EXCLUDED.unit, role = EXCLUDED.role, is_main = EXCLUDED.is_main, match_weight = EXCLUDED.match_weight;")
    lines.append("")

    lines.append("INSERT INTO recipe_tags (id, recipe_id, tag, tag_type) VALUES")
    tag_values = [
        f"  ({sql(stable_uuid('recipe_tag', str(row['id'])))}, {sql(stable_uuid('recipe', str(row['recipe_id'])))}, {sql(str(row['tag']))}, {sql(str(row.get('tag_type') or 'attribute'))})"
        for row in recipe_tags
    ]
    lines.append(",\n".join(tag_values) + "\nON CONFLICT (recipe_id, tag) DO UPDATE SET tag_type = EXCLUDED.tag_type;")
    lines.append("")

    lines.append("INSERT INTO recipe_pairings (id, recipe_id, paired_recipe_id, pairing_message, pairing_strength) VALUES")
    pairing_values = []
    for row in recipe_pairings:
        paired = row.get("paired_recipe_id")
        paired_recipe_sql = (
            sql(stable_uuid("recipe", str(paired)))
            if paired and str(paired) in recipe_names
            else "NULL"
        )
        pairing_values.append(
            "  ("
            + ", ".join(
                [
                    sql(stable_uuid("recipe_pairing", str(row["id"]))),
                    sql(stable_uuid("recipe", str(row["recipe_id"]))),
                    paired_recipe_sql,
                    sql(str(row["pairing_message"])),
                    sql(row.get("pairing_strength") or 0.85),
                ]
            )
            + ")"
        )
    lines.append(",\n".join(pairing_values) + "\nON CONFLICT (id) DO UPDATE SET paired_recipe_id = EXCLUDED.paired_recipe_id, pairing_message = EXCLUDED.pairing_message, pairing_strength = EXCLUDED.pairing_strength;")
    lines.append("")

    lines.append("INSERT INTO ingredient_aliases (id, ingredient_id, alias_name) VALUES")
    alias_values = [
        f"  ({sql(stable_uuid('ingredient_alias', str(row['id'])))}, {sql(stable_uuid('ingredient', str(row['ingredient_id'])))}, {sql(str(row['alias_name']))})"
        for row in ingredient_aliases
    ]
    lines.append(",\n".join(alias_values) + "\nON CONFLICT (ingredient_id, alias_name) DO NOTHING;")
    lines.append("")

    lines.append("COMMIT;")
    return "\n".join(lines) + "\n"


def main() -> None:
    if not MAPPING_XLSX.exists():
        raise SystemExit(f"Missing source workbook: {MAPPING_XLSX}")
    mapping = load_mapping()
    OUTPUT_SQL.write_text(build_seed(mapping), encoding="utf-8")
    OUTPUT_JSON.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_JSON.write_text(json.dumps(build_fallback_json(mapping), indent=2), encoding="utf-8")
    print(f"Wrote {OUTPUT_SQL}")
    print(f"Wrote {OUTPUT_JSON}")


if __name__ == "__main__":
    main()
